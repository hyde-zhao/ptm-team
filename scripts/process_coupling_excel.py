"""
加工 .input/ 中的 Excel 耦合矩阵为结构化 YAML。

输入: .input/NGFW系列 V60R1C00 特性树&耦合矩阵&形态差异三表.xlsx
输出:
  - resource/coupling-matrix/tgfw-feature-tree.yaml   — 特性分层树
  - resource/coupling-matrix/tgfw-coupling-matrix.yaml — 合并耦合矩阵
  - resource/coupling-matrix/tgfw-platform-diff.yaml   — 形态差异
"""

import openpyxl
import yaml
import os
import re
from pathlib import Path

ROOT = Path("/home/hyde/projects/ptm-team")
EXCEL_PATH = ROOT / ".input/NGFW系列 V60R1C00 特性树&耦合矩阵&形态差异三表.xlsx"
OUT_DIR = ROOT / "resource/coupling-matrix"


def clean_header(text):
    """清理列标题中的换行符和常见问题。"""
    if text is None:
        return None
    text = str(text).replace("\n", "").strip()
    # 修正常见 OCR/格式问题
    text = text.replace("bvI", "bvi").replace("ByPASS", "Bypass")
    text = text.replace("byPass", "Bypass")
    # 移除多余空格
    text = re.sub(r"\s+", "", text)
    return text


def process_feature_tree(wb):
    """从「V60R1C00版本特性树」提取分层特性树，补齐合并单元格。"""
    ws = wb["V60R1C00版本特性树"]
    features = []

    # 层级缓存——Excel 使用合并单元格，需向下填充
    l1_cache = l2_cache = l3_cache = l4_cache = l5_cache = None

    for row_idx in range(3, ws.max_row + 1):
        v1 = ws.cell(row=row_idx, column=1).value
        v2 = ws.cell(row=row_idx, column=2).value
        v3 = ws.cell(row=row_idx, column=3).value
        v4 = ws.cell(row=row_idx, column=4).value
        v5 = ws.cell(row=row_idx, column=5).value

        # 向下填充合并单元格
        if v1 is not None:
            l1_cache = str(v1).strip()
            l2_cache = l3_cache = l4_cache = l5_cache = None  # 新一级目录重置下级
        if v2 is not None:
            l2_cache = str(v2).strip()
            l3_cache = l4_cache = l5_cache = None
        if v3 is not None:
            l3_cache = str(v3).strip()
            l4_cache = l5_cache = None
        if v4 is not None:
            l4_cache = str(v4).strip()
            l5_cache = None
        if v5 is not None:
            l5_cache = str(v5).strip()

        row_data = {
            "level1": l1_cache,
            "level2": l2_cache,
            "level3": l3_cache,
            "level4": l4_cache,
            "level5": l5_cache,
            "definition": ws.cell(row=row_idx, column=6).value,
            "name_en": ws.cell(row=row_idx, column=7).value,
            "feature_id": ws.cell(row=row_idx, column=8).value,
            "version_change": ws.cell(row=row_idx, column=9).value,
            "risk_level": ws.cell(row=row_idx, column=11).value,
            "priority": ws.cell(row=row_idx, column=15).value,
        }

        # 跳过全空行
        if all(v is None for v in row_data.values()):
            continue

        # 清理字符串
        for k in ["level1", "level2", "level3", "level4", "level5",
                   "definition", "name_en", "version_change", "risk_level"]:
            if isinstance(row_data[k], str):
                row_data[k] = row_data[k].strip()

        # 将 特性ID 转为字符串（避免科学计数法）
        if row_data["feature_id"] is not None:
            row_data["feature_id"] = str(int(row_data["feature_id"]))

        features.append(row_data)

    return features


def process_coupling_matrix_sheet(ws, domain):
    """处理一个耦合矩阵 Sheet，返回 coupling 记录列表。"""
    # Row 1: 图例行 (强耦合/弱耦合/不耦合)
    # Row 2: 列标题行 — 特性名列表
    # Row 3+: 数据行

    # 提取列标题 (col 3-63, 即 0-indexed 2-62)
    col_headers = []
    for col in range(2, ws.max_column):  # 0-indexed: col 2 = C 列
        val = clean_header(ws.cell(row=2, column=col + 1).value)  # row=2 是第 2 行
        if val:
            col_headers.append(val)

    # 重新读取：openpyxl 使用 1-indexed
    col_headers = []
    for col in range(3, ws.max_column + 1):  # col 3 = C 列
        val = clean_header(ws.cell(row=2, column=col).value)
        if val:
            col_headers.append(val)

    print(f"  [{domain}] 列标题数: {len(col_headers)}")

    # 遍历数据行
    couplings = []
    current_l3 = None  # 当前三级目录

    for row_idx in range(3, ws.max_row + 1):
        l3_raw = ws.cell(row=row_idx, column=1).value   # A 列
        l4_raw = ws.cell(row=row_idx, column=2).value   # B 列

        # 更新当前三级目录
        if l3_raw is not None:
            current_l3 = str(l3_raw).strip()

        l4 = str(l4_raw).strip() if l4_raw is not None else None

        # 构建源特性全路径
        if l4 and current_l3:
            source_feature = f"{current_l3} > {l4}"
        elif current_l3:
            source_feature = current_l3
        else:
            continue  # 无有效行标签

        # 扫描耦合标记
        for col_idx in range(3, ws.max_column + 1):
            cell_val = ws.cell(row=row_idx, column=col_idx).value
            if cell_val is not None:
                cell_str = str(cell_val).strip()
                if cell_str == "√":
                    target_idx = col_idx - 3  # col 3 → index 0
                    if target_idx < len(col_headers):
                        target_feature = col_headers[target_idx]
                        couplings.append({
                            "source_feature": source_feature,
                            "target_feature": target_feature,
                            "coupling_level": "strong",
                            "domain": domain,
                        })

    return couplings, current_l3  # current_l3 for logging


def process_platform_diff(wb):
    """从「形态差异」Sheet 提取平台差异矩阵。"""
    ws = wb["形态差异"]
    # Row 1: headers (三级目录/四级目录/责任人/X86系列/ARM系列)
    # Row 2: sub-headers (J1900/C236/EP/待定/待定/待定)
    # Row 3+: data — 目前平台列数据尚未填入

    platform_cols = {
        4: "J1900",
        5: "C236",
        6: "EP",
        7: "ARM_待定1",
        8: "ARM_待定2",
        9: "ARM_待定3",
    }

    features = []  # 特性列表（骨架）
    diffs = []     # 有平台数据的记录
    current_l3 = None

    for row_idx in range(3, ws.max_row + 1):
        l3_raw = ws.cell(row=row_idx, column=1).value
        l4_raw = ws.cell(row=row_idx, column=2).value

        if l3_raw is not None:
            current_l3 = str(l3_raw).strip()

        l4 = str(l4_raw).strip() if l4_raw is not None else None

        if l4 and current_l3:
            source_feature = f"{current_l3} > {l4}"
        elif current_l3:
            source_feature = current_l3
        else:
            continue

        features.append(source_feature)

        platforms = {}
        for col_idx, plat_name in platform_cols.items():
            cell_val = ws.cell(row=row_idx, column=col_idx).value
            if cell_val is not None:
                platforms[plat_name] = str(cell_val).strip()

        if platforms:
            diffs.append({
                "feature": source_feature,
                "platforms": platforms,
            })

    return features, diffs


def main():
    print(f"读取 Excel: {EXCEL_PATH}")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    print(f"Sheet 列表: {wb.sheetnames}")

    # 1. 特性树
    print("\n=== 处理特性树 ===")
    features = process_feature_tree(wb)
    print(f"  提取 {len(features)} 条特性记录")

    with open(OUT_DIR / "tgfw-feature-tree.yaml", "w", encoding="utf-8") as f:
        yaml.dump(
            {"schema_version": 1, "product": "TGFW", "feature_tree": features},
            f, allow_unicode=True, default_flow_style=False, sort_keys=False,
            width=200,
        )
    print(f"  输出: tgfw-feature-tree.yaml")

    # 2. 耦合矩阵（合并两个 Sheet）
    print("\n=== 处理耦合矩阵 ===")
    all_couplings = []

    for sheet_name, domain in [
        ("耦合矩阵（数通&高可用）", "datacom-ha"),
        ("耦合矩阵（安全&系统）", "security-system"),
    ]:
        ws = wb[sheet_name]
        couplings, _ = process_coupling_matrix_sheet(ws, domain)
        print(f"  [{domain}] 耦合关系数: {len(couplings)}")
        all_couplings.extend(couplings)

    print(f"  总计耦合关系数: {len(all_couplings)}")

    with open(OUT_DIR / "tgfw-coupling-matrix.yaml", "w", encoding="utf-8") as f:
        # 写入自定义头部
        f.write(f"# TGFW 产品耦合矩阵（LLM 可读格式）\n")
        f.write(f"# 合并自 Excel 两个 Sheet：数通&高可用 + 安全&系统\n")
        f.write(f"# coupling_level: strong = 强耦合，标记为 √\n")
        f.write(f"# domain: datacom-ha | security-system\n")
        f.write(f"# 总记录数: {len(all_couplings)}\n")
        yaml.dump(
            {"schema_version": 1, "product": "TGFW",
             "total_couplings": len(all_couplings),
             "couplings": all_couplings},
            f, allow_unicode=True, default_flow_style=False, sort_keys=False,
            width=200,
        )
    print(f"  输出: tgfw-coupling-matrix.yaml")

    # 3. 形态差异
    print("\n=== 处理形态差异 ===")
    features, diffs = process_platform_diff(wb)
    print(f"  特性条目数: {len(features)}")
    print(f"  有平台数据的记录: {len(diffs)}")

    status = "data_available" if diffs else "empty"
    note = "" if diffs else "平台差异数据尚未填入，当前仅包含特性结构骨架"

    with open(OUT_DIR / "tgfw-platform-diff.yaml", "w", encoding="utf-8") as f:
        f.write(f"# TGFW 形态差异矩阵（LLM 可读格式）\n")
        if note:
            f.write(f"# ⚠ {note}\n")
        yaml.dump(
            {"schema_version": 1, "product": "TGFW",
             "platforms": ["J1900", "C236", "EP", "ARM"],
             "status": status,
             "note": note,
             "total_features": len(features),
             "total_with_data": len(diffs),
             "features": features,
             "diffs": diffs},
            f, allow_unicode=True, default_flow_style=False, sort_keys=False,
            width=200,
        )
    print(f"  输出: tgfw-platform-diff.yaml")

    print("\n✅ 全部完成")


if __name__ == "__main__":
    main()
